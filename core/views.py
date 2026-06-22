from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.utils import timezone
from django.db.models import Count, Avg, Q, Sum
from functools import wraps
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .logros import verificar_logros
from django.db.models import Count, Sum
from .models import SesionEntrenamiento, EjercicioSesion, UsuarioLogro
from django.contrib.auth.hashers import (
    make_password,
    check_password
)
from django.conf import settings
from django.http import FileResponse, Http404


import openpyxl
import os

from .models import (
    Usuario, TipoUsuario, Ejercicio, TipoEjercicio, Rutina,
    RutinaEjercicio, RutinaUsuario, RutinaUsuarioDia, Seguimiento, Logro,
    UsuarioLogro, Notificacion, ValoracionProfesional, Nivel, Medida,
    SesionEntrenamiento, EjercicioSesion,
    VideoEjercicio
)

from .utils import (
    enviar_email_bienvenida,
    buscar_videos_youtube,
    crear_notificacion
)
from .importacion import importar_ejercicios, importar_usuarios, importar_rutinas
from functools import wraps
from .decorators import rol_required
from django.shortcuts import render
from django.db.models import Avg
from .models import Seguimiento, Rutina, Ejercicio, Usuario 
from django.db.models import Count






 # ─── DESCARGAR REPORTES ─────────────────

@rol_required('Admin')
def exportar_reportes_excel(request):
    wb = openpyxl.Workbook()

    # ─── HOJA 1: RESUMEN ─────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    ws1.append(["Usuarios", "Coaches", "Ejercicios", "Rutinas", "Sesiones"])

    ws1.append([
        Usuario.objects.filter(id_tipo_usuario__rol='Usuario').count(),
        Usuario.objects.filter(id_tipo_usuario__rol='Coach').count(),
        Ejercicio.objects.count(),
        Rutina.objects.count(),
        Seguimiento.objects.count()
    ])

    # ─── HOJA 2: USUARIOS POR GENERO ─────────────────
    ws2 = wb.create_sheet("Usuarios por género")

    ws2.append(["Género", "Total"])
    datos_genero = (
        Usuario.objects.filter(id_tipo_usuario__rol='Usuario')
        .values('genero')
        .annotate(total=Count('id_usuario'))
    )

    for g in datos_genero:
        ws2.append([g['genero'], g['total']])

    # ─── HOJA 3: EJERCICIOS POR NIVEL ────────────────
    ws3 = wb.create_sheet("Ejercicios por nivel")

    ws3.append(["Nivel", "Total"])
    datos_nivel = (
        Ejercicio.objects.values('nivel_dificultad')
        .annotate(total=Count('id_ejercicio'))
    )

    for e in datos_nivel:
        ws3.append([e['nivel_dificultad'], e['total']])

    # ─── HOJA 4: TOP EJERCICIOS ──────────────────────
    ws4 = wb.create_sheet("Top ejercicios")

    ws4.append(["Ejercicio", "Total"])
    top = (
        Seguimiento.objects.values('id_ejercicio__nombre')
        .annotate(total=Count('id_progreso'))
        .order_by('-total')[:5]
    )

    for t in top:
        ws4.append([t['id_ejercicio__nombre'], t['total']])

    # ─── HOJA 5: RUTINAS MAS USADAS ─────────────────
    ws5 = wb.create_sheet("Rutinas")

    ws5.append(["Rutina", "Nivel", "Disciplina", "Asignaciones"])

    rutinas = (
        Rutina.objects.select_related('id_nivel')
        .annotate(total=Count('rutinausuario'))
        .order_by('-total')[:10]
    )

    for r in rutinas:
        ws5.append([
            r.nombre,
            r.id_nivel.nombre if r.id_nivel else '',
            r.disciplina,
            r.total
        ])

    # ─── RESPUESTA ──────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_completo.xlsx'

    wb.save(response)
    return response


# ─── Decoradores de autenticación ────────────────────────────────────────────

def login_required_custom(f):
    @wraps(f)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')
        return f(request, *args, **kwargs)
    return wrapper


def rol_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(request, *args, **kwargs):
            if 'usuario_id' not in request.session:
                return redirect('login')
            if request.session.get('rol') not in roles:
                messages.error(request, 'No tienes permiso para acceder a esta sección.')
                return redirect('dashboard')
            return f(request, *args, **kwargs)
        return wrapper
    return decorator


def get_usuario_sesion(request):
    uid = request.session.get('usuario_id')
    if uid:
        try:
            return Usuario.objects.select_related('id_tipo_usuario').get(pk=uid)
        except Usuario.DoesNotExist:
            pass
    return None


def notif_count(request):
    
    if 'usuario_id' not in request.session:
        return 0

    return Notificacion.objects.filter(
        id_usuario_id=request.session['usuario_id'],
        estado='No Leída'
    ).count()


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if 'usuario_id' in request.session:
        return redirect('dashboard')

    mostrar_registro = False   # 👈 bandera para el template

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        # ───────── LOGIN ─────────
        if form_type == 'login':
            username = request.POST.get('usuario', '').strip()
            password = request.POST.get('contrasena', '')

            try:
                user = Usuario.objects.select_related('id_tipo_usuario').get(usuario=username)

                if check_password(password, user.contrasena):
                    request.session['usuario_id'] = user.id_usuario
                    request.session['rol'] = user.id_tipo_usuario.rol
                    request.session['nombre'] = f"{user.nombre} {user.apellido}"

                    messages.success(request, f'¡Bienvenido, {user.nombre}!')
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Usuario o contraseña incorrectos.')

            except Usuario.DoesNotExist:
                messages.error(request, 'Usuario o contraseña incorrectos.')

        # ───────── REGISTRO ─────────
        elif form_type == 'registro':
            mostrar_registro = True   # 👈 abrir pestaña registro si falla algo

            nombre = request.POST.get('nombre')
            apellido = request.POST.get('apellido')
            usuario = request.POST.get('usuario')
            email = request.POST.get('email')
            genero = request.POST.get('genero')
            password = request.POST.get('contrasena')
            password2 = request.POST.get('contrasena2')

            # Validaciones
            if password != password2:
                messages.error(request, 'Las contraseñas no coinciden.')
                return render(request, 'core/login.html', {'mostrar_registro': True})

            if Usuario.objects.filter(usuario=usuario).exists():
                messages.error(request, 'El usuario ya existe.')
                return render(request, 'core/login.html', {'mostrar_registro': True})

            # crear usuario tipo "Usuario"
            tipo_usuario = TipoUsuario.objects.get(rol="Usuario")

            nuevo = Usuario.objects.create(
                nombre=nombre,
                apellido=apellido,
                usuario=usuario,
                email=email,
                genero=genero,
                id_tipo_usuario=tipo_usuario,
                contrasena=make_password(password)
            )

            messages.success(request, 'Cuenta creada correctamente. Ahora puedes iniciar sesión.')
            return redirect('login')

    return render(request, 'core/login.html', {'mostrar_registro': mostrar_registro})

def registro_view(request):
    print("ENTRÓ A REGISTRO_VIEW")

    if 'usuario_id' in request.session:
        return redirect('dashboard')

    if request.method == 'POST':
        print("FORMULARIO ENVIADO")

        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        username = request.POST.get('usuario', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('contrasena', '')
        password2 = request.POST.get('contrasena2', '')
        genero = request.POST.get('genero', 'Otro')
        edad = request.POST.get('edad') or None

        print("Datos recibidos:")
        print(nombre, apellido, username, email, genero)

        if not all([nombre, apellido, username, email, password]):
            print("FALTAN CAMPOS")
            messages.error(request, 'Todos los campos obligatorios deben completarse.')

        elif password != password2:
            print("CONTRASEÑAS DISTINTAS")
            messages.error(request, 'Las contraseñas no coinciden.')

        elif Usuario.objects.filter(usuario=username).exists():
            print("USUARIO YA EXISTE")
            messages.error(request, 'Ese nombre de usuario ya está en uso.')

        elif Usuario.objects.filter(email=email).exists():
            print("EMAIL YA EXISTE")
            messages.error(request, 'Ese correo ya está registrado.')

        else:
            try:
                print("CREANDO USUARIO...")

                tipo, creado = TipoUsuario.objects.get_or_create(rol='Usuario')

                Usuario.objects.create(
                    nombre=nombre,
                    apellido=apellido,
                    usuario=username,
                    email=email,
                    contrasena=make_password(password),
                    genero=genero,
                    edad=edad,
                    id_tipo_usuario=tipo,
                    fecha_registro=timezone.now().date()
                )

                print("USUARIO CREADO")

                messages.success(request, '¡Cuenta creada! Ya puedes iniciar sesión.')
                enviar_email_bienvenida(nombre, email)

                return redirect('login')

            except Exception as e:
                print("ERROR:", e)
                messages.error(request, f'Error interno: {e}')

    return render(request, 'core/registro.html')




def logout_view(request):
    request.session.flush()
    return redirect('login')


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required_custom
def dashboard_view(request):
    rol = request.session.get('rol')
    uid = request.session.get('usuario_id')
    ctx = {'notif_count': notif_count(request), 'rol': rol}

    if rol == 'Admin':
        ctx['total_usuarios'] = Usuario.objects.count()
        ctx['total_ejercicios'] = Ejercicio.objects.count()
        ctx['total_rutinas'] = Rutina.objects.count()
        ctx['total_coaches'] = Usuario.objects.filter(id_tipo_usuario__rol='Coach').count()
        ctx['ultimos_usuarios'] = Usuario.objects.select_related('id_tipo_usuario').order_by('-fecha_registro')[:5]

    elif rol == 'Coach':
        ctx['total_rutinas'] = Rutina.objects.count()
        ctx['total_ejercicios'] = Ejercicio.objects.count()
        ctx['total_usuarios_asignados'] = RutinaUsuario.objects.filter(estado='Activa').values('id_usuario').distinct().count()
        ctx['rutinas_activas'] = RutinaUsuario.objects.filter(estado='Activa').select_related('id_usuario', 'id_rutina')[:5]

    elif rol == 'Usuario':
        ctx['mis_rutinas'] = RutinaUsuario.objects.filter(
            id_usuario_id=uid, estado='Activa'
        ).select_related('id_rutina')[:3]
        ctx['mis_logros'] = UsuarioLogro.objects.filter(
            id_usuario_id=uid
        ).select_related('id_logro')[:4]
        ctx['total_sesiones'] = Seguimiento.objects.filter(id_usuario_id=uid).count()
        ctx['notificaciones'] = Notificacion.objects.filter(
            id_usuario_id=uid, estado='No Leída'
        ).order_by('-fecha_hora')[:3]

    return render(request, 'core/dashboard.html', ctx)



# ─── Finalizar una rutina ─────────────────────────────────────

@login_required_custom
@rol_required('Coach', 'Admin')
def coach_finalizar_asignacion(request, pk):

    asignacion = get_object_or_404(
        RutinaUsuario,
        pk=pk
    )

    asignacion.estado = 'Finalizada'
    asignacion.fecha_final = timezone.now().date()
    asignacion.save()

    messages.success(
        request,
        'La rutina fue desasignada correctamente.'
    )

    return redirect('coach_usuarios')

# ─── ADMIN: Usuarios ──────────────────────────────────────────────────────────

@rol_required('Admin')
def admin_usuarios(request):
    q = request.GET.get('q', '')
    rol_f = request.GET.get('rol', '')
    usuarios = Usuario.objects.select_related('id_tipo_usuario')
    if q:
        usuarios = usuarios.filter(nombre__icontains=q) | usuarios.filter(apellido__icontains=q) | usuarios.filter(usuario__icontains=q)
    if rol_f:
        usuarios = usuarios.filter(id_tipo_usuario__rol=rol_f)
    return render(request, 'core/admin/usuarios.html', {
        'usuarios': usuarios.order_by('nombre'),
        'q': q, 'rol_f': rol_f,
        'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_crear_usuario(request):
    tipos = TipoUsuario.objects.all()
    if request.method == 'POST':
        try:
            tipo = TipoUsuario.objects.get(pk=request.POST['id_tipo_usuario'])
            Usuario.objects.create(
                nombre=request.POST['nombre'],
                apellido=request.POST['apellido'],
                usuario=request.POST['usuario'],
                email=request.POST['email'],
                contrasena=make_password(request.POST['contrasena']),
                genero=request.POST.get('genero', 'Otro'),
                edad=request.POST.get('edad') or None,
                objetivo=request.POST.get('objetivo', ''),
                id_tipo_usuario=tipo,
                fecha_registro=timezone.now().date()
            )
            messages.success(request, 'Usuario creado correctamente.')
            return redirect('admin_usuarios')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/admin/usuario_form.html', {
        'tipos': tipos, 'accion': 'Crear', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_editar_usuario(request, pk):
    user = get_object_or_404(Usuario, pk=pk)
    tipos = TipoUsuario.objects.all()
    if request.method == 'POST':
        try:
            user.nombre = request.POST['nombre']
            user.apellido = request.POST['apellido']
            user.usuario = request.POST['usuario']
            user.email = request.POST['email']
            user.genero = request.POST.get('genero', user.genero)
            user.edad = request.POST.get('edad') or None
            user.objetivo = request.POST.get('objetivo', '')
            user.id_tipo_usuario = TipoUsuario.objects.get(pk=request.POST['id_tipo_usuario'])
            if request.POST.get('contrasena'):
                user.contrasena = make_password(request.POST['contrasena'])
            user.save()
            messages.success(request, 'Usuario actualizado.')
            return redirect('admin_usuarios')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/admin/usuario_form.html', {
        'user_obj': user, 'tipos': tipos, 'accion': 'Editar', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_eliminar_usuario(request, pk):
    user = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Usuario eliminado.')
        return redirect('admin_usuarios')
    return render(request, 'core/admin/confirmar_eliminar.html', {
        'objeto': user, 'tipo': 'usuario', 'notif_count': notif_count(request)
    })


@login_required_custom
@rol_required('Admin', 'Coach')
def eliminar_rutina_ejercicio(request, pk):

    relacion = get_object_or_404(
        RutinaEjercicio,
        pk=pk
    )

    rutina_id = relacion.id_rutina.id_rutina

    if request.method == 'POST':

        relacion.delete()

        messages.success(
            request,
            'Ejercicio eliminado de la rutina.'
        )

        return redirect(
            'ver_rutina',
            rutina_id=rutina_id
        )

    return render(
        request,
        'core/coach/confirmar_eliminar_rutina_ejercicio.html',
        {
            'relacion': relacion,
            'notif_count': notif_count(request)
        }
    )


@login_required_custom
@rol_required('Admin', 'Coach')
def editar_rutina_ejercicio(request, pk):

    relacion = get_object_or_404(
        RutinaEjercicio,
        pk=pk
    )

    if request.method == 'POST':

        relacion.orden = request.POST.get('orden') or None
        relacion.series = request.POST.get('series') or None
        relacion.repeticiones = request.POST.get('repeticiones') or None
        relacion.descanso = request.POST.get('descanso')

        relacion.save()

        messages.success(
            request,
            'Ejercicio actualizado.'
        )

        return redirect(
            'ver_rutina',
            rutina_id=relacion.id_rutina.id_rutina
        )

    return render(
        request,
        'core/coach/editar_rutina_ejercicio.html',
        {
            'relacion': relacion,
            'notif_count': notif_count(request)
        }
    )


@login_required_custom
@rol_required('Admin', 'Coach')
def editar_asignacion_rutina(request, pk):

    asignacion = get_object_or_404(
        RutinaUsuario,
        pk=pk
    )

    if request.method == 'POST':

        asignacion.fecha_inicio = (
            request.POST.get('fecha_inicio') or None
        )

        asignacion.fecha_final = (
            request.POST.get('fecha_final') or None
        )

        asignacion.adaptaciones_personalizadas = (
            request.POST.get('adaptaciones', '')
        )

        asignacion.save()

        dias_seleccionados = request.POST.getlist('dias')

        print("DIAS RECIBIDOS:", dias_seleccionados)

        # borrar días anteriores
        RutinaUsuarioDia.objects.filter(
            id_rutina_usuario=asignacion
        ).delete()

        # crear nuevos días
        for dia in dias_seleccionados:

            RutinaUsuarioDia.objects.create(
                id_rutina_usuario=asignacion,
                dia_semana=dia
            )

        print(
            "DIAS GUARDADOS:",
            list(
                RutinaUsuarioDia.objects.filter(
                    id_rutina_usuario=asignacion
                ).values_list(
                    'dia_semana',
                    flat=True
                )
            )
        )

        messages.success(
            request,
            'Asignación actualizada correctamente.'
        )

        return redirect('coach_usuarios')

    dias_actuales = list(
        RutinaUsuarioDia.objects.filter(
            id_rutina_usuario=asignacion
        ).values_list(
            'dia_semana',
            flat=True
        )
    )

    print("ASIGNACION:", asignacion.pk)
    print("DIAS ACTUALES:", dias_actuales)

    return render(
        request,
        'core/coach/editar_asignacion_rutina.html',
        {
            'asignacion': asignacion,
            'dias_actuales': dias_actuales,
            'notif_count': notif_count(request)
        }
    )








@login_required_custom
@rol_required('Admin', 'Coach')
def editar_planificacion_rutina(request, rutina_id):

    rutina = get_object_or_404(
        Rutina,
        pk=rutina_id
    )

    ejercicios = RutinaEjercicio.objects.filter(
        id_rutina=rutina
    ).select_related(
        'id_ejercicio'
    ).order_by(
        'dia_semana',
        'orden'
    )

    dias = {
        'Lunes': [],
        'Martes': [],
        'Miércoles': [],
        'Jueves': [],
        'Viernes': [],
        'Sábado': [],
        'Domingo': []
    }

    for ejercicio in ejercicios:
        dias[ejercicio.dia_semana].append(ejercicio)

    return render(
        request,
        'core/coach/editar_planificacion_rutina.html',
        {
            'rutina': rutina,
            'dias': dias,
            'notif_count': notif_count(request)
        }
    )

@login_required_custom
@rol_required('Admin', 'Coach')
def mover_ejercicio_dia(request, pk):

    relacion = get_object_or_404(
        RutinaEjercicio,
        pk=pk
    )

    if request.method == 'POST':

        nuevo_dia = request.POST.get('dia_semana')

        dias_validos = [
            'Lunes',
            'Martes',
            'Miércoles',
            'Jueves',
            'Viernes',
            'Sábado',
            'Domingo'
        ]

        if nuevo_dia in dias_validos:

            relacion.dia_semana = nuevo_dia
            relacion.save()

            messages.success(
                request,
                'Ejercicio movido correctamente.'
            )

    return redirect(
        'editar_planificacion_rutina',
        rutina_id=relacion.id_rutina.id_rutina
    )


@rol_required('Admin')
def admin_cambiar_rol(request, pk):
    if request.method == 'POST':
        user = get_object_or_404(Usuario, pk=pk)
        nuevo_rol = request.POST.get('nuevo_rol')
        if nuevo_rol not in ['Usuario', 'Coach', 'Admin']:
            messages.error(request, 'Rol inválido.')
        else:
            try:
                tipo = TipoUsuario.objects.get(rol=nuevo_rol)
                user.id_tipo_usuario = tipo
                user.save()
                messages.success(request, f'El rol de {user.nombre} fue cambiado a {nuevo_rol}.')
            except TipoUsuario.DoesNotExist:
                messages.error(request, f'No existe el rol {nuevo_rol} en el sistema.')
    return redirect('admin_usuarios')


# ─── ADMIN: Ejercicios ────────────────────────────────────────────────────────

from django.http import HttpResponse



@rol_required('Admin')

def admin_ejercicios(request):

    q = request.GET.get('q', '')



    ejercicios = Ejercicio.objects.select_related(

        'tipo_ejercicio'

    )



    if q:

        ejercicios = ejercicios.filter(

            nombre__icontains=q

        )



    return render(request, 'core/admin/ejercicios.html', {

        'ejercicios': ejercicios.order_by('nombre'),

        'q': q,

        'notif_count': notif_count(request)

    }) 





@rol_required('Admin')
def admin_crear_ejercicio(request):
    tipos = TipoEjercicio.objects.all()

    if request.method == 'POST':
        try:
            ejercicio = Ejercicio.objects.create(
                nombre=request.POST['nombre'],
                tipo_ejercicio=TipoEjercicio.objects.get(pk=request.POST['tipo_ejercicio']) if request.POST.get('tipo_ejercicio') else None,
                equipo_necesario=request.POST.get('equipo_necesario', ''),
                nivel_dificultad=request.POST['nivel_dificultad'],
                instrucciones=request.POST.get('instrucciones', ''),
            )

            # 🔥 GUARDAR VIDEOS
            for i in range(1, 4):
                url = request.POST.get(f'url_video_{i}')
                if url:
                    VideoEjercicio.objects.create(
                        ejercicio=ejercicio,
                        url=url
                    )

            messages.success(request, 'Ejercicio creado con videos.')
            return redirect('admin_ejercicios')  # 👈 IMPORTANTE (no coach)

        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'core/admin/ejercicio_form.html', {  # 👈 ruta correcta
        'tipos': tipos,
        'accion': 'Crear',
        'notif_count': notif_count(request)
    })


@rol_required('Admin', 'Coach')
def coach_editar_ejercicio(request, pk):
    ej = get_object_or_404(Ejercicio, pk=pk)
    tipos = TipoEjercicio.objects.all()

    # 🔥 obtener videos actuales desde BD
    videos_db = list(ej.videos.all())
    urls = ["", "", ""]

    for i in range(min(3, len(videos_db))):
        urls[i] = videos_db[i].url

    # ==============================
    # POST → GUARDAR CAMBIOS
    # ==============================
    if request.method == 'POST':
        try:
            ej.nombre = request.POST['nombre']

            tipo_id = request.POST.get('tipo_ejercicio')
            ej.tipo_ejercicio = TipoEjercicio.objects.get(id_tipo=tipo_id) if tipo_id else None

            ej.equipo_necesario = request.POST.get('equipo_necesario', '')
            ej.nivel_dificultad = request.POST['nivel_dificultad']
            ej.instrucciones = request.POST.get('instrucciones', '')

            ej.save()

            # 🔥 BORRAR videos anteriores
            ej.videos.all().delete()

            # 🔥 GUARDAR nuevos videos
            for i in range(1, 4):
                url = request.POST.get(f'url_video_{i}')
                if url:
                    VideoEjercicio.objects.create(
                        ejercicio=ej,
                        url=url
                    )

            messages.success(request, 'Ejercicio actualizado correctamente.')
            return redirect('coach_ejercicios')

        except Exception as e:
            messages.error(request, f'Error: {e}')


    # ==============================
    # GET → CARGAR FORMULARIO
    # ==============================
    return render(request, 'core/ejercicios/ejercicio_form.html', {
        'obj': ej,         # 👈 usa obj (tu html ya lo usa)
        'tipos': tipos,
        'urls': urls,      # 👈 IMPORTANTE
        'accion': 'Editar',
        'notif_count': notif_count(request)
    })




@rol_required('Admin')
def admin_eliminar_ejercicio(request, pk):
    ej = get_object_or_404(Ejercicio, pk=pk)
    if request.method == 'POST':
        ej.delete()
        messages.success(request, 'Ejercicio eliminado.')
        return redirect('admin_ejercicios')
    return render(request, 'core/admin/confirmar_eliminar.html', {
        'objeto': ej, 'tipo': 'ejercicio', 'notif_count': notif_count(request)
    })


# ─── ADMIN: Rutinas ───────────────────────────────────────────────────────────

@rol_required('Admin')
def admin_rutinas(request):
    rutinas = Rutina.objects.select_related('id_nivel').order_by('nombre')
    return render(request, 'core/admin/rutinas.html', {
        'rutinas': rutinas, 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_crear_rutina(request):
    niveles = Nivel.objects.all()
    if request.method == 'POST':
        try:
            Rutina.objects.create(
                nombre=request.POST['nombre'],
                descripcion=request.POST.get('descripcion', ''),
                id_nivel=Nivel.objects.get(pk=request.POST['id_nivel']) if request.POST.get('id_nivel') else None,
                disciplina=request.POST.get('disciplina', ''),
                duracion_total=request.POST.get('duracion_total', ''),
                comentario=request.POST.get('comentario', ''),
            )
            messages.success(request, 'Rutina creada.')
            return redirect('admin_rutinas')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/admin/rutina_form.html', {
        'niveles': niveles, 'accion': 'Crear', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_editar_rutina(request, pk):
    r = get_object_or_404(Rutina, pk=pk)
    niveles = Nivel.objects.all()
    if request.method == 'POST':
        try:
            r.nombre = request.POST['nombre']
            r.descripcion = request.POST.get('descripcion', '')
            r.id_nivel = Nivel.objects.get(pk=request.POST['id_nivel']) if request.POST.get('id_nivel') else None
            r.disciplina = request.POST.get('disciplina', '')
            r.duracion_total = request.POST.get('duracion_total', '')
            r.comentario = request.POST.get('comentario', '')
            r.save()
            messages.success(request, 'Rutina actualizada.')
            return redirect('admin_rutinas')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/admin/rutina_form.html', {
        'obj': r, 'niveles': niveles, 'accion': 'Editar', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_eliminar_rutina(request, pk):
    r = get_object_or_404(Rutina, pk=pk)
    if request.method == 'POST':
        r.delete()
        messages.success(request, 'Rutina eliminada.')
        return redirect('admin_rutinas')
    return render(request, 'core/admin/confirmar_eliminar.html', {
        'objeto': r, 'tipo': 'rutina', 'notif_count': notif_count(request)
    })


# ─── ADMIN: Tipos de Ejercicio ────────────────────────────────────────────────

@rol_required('Admin')
def admin_tipos_ejercicio(request):
    tipos = TipoEjercicio.objects.all()
    return render(request, 'core/admin/tipos_ejercicio.html', {
        'tipos': tipos, 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_crear_tipo_ejercicio(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if nombre:
            TipoEjercicio.objects.create(nombre=nombre)
            messages.success(request, 'Tipo creado.')
            return redirect('admin_tipos_ejercicio')
    return render(request, 'core/admin/tipo_ejercicio_form.html', {
        'accion': 'Crear', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_editar_tipo_ejercicio(request, pk):
    t = get_object_or_404(TipoEjercicio, pk=pk)
    if request.method == 'POST':
        t.nombre = request.POST.get('nombre', t.nombre)
        t.save()
        messages.success(request, 'Tipo actualizado.')
        return redirect('admin_tipos_ejercicio')
    return render(request, 'core/admin/tipo_ejercicio_form.html', {
        'obj': t, 'accion': 'Editar', 'notif_count': notif_count(request)
    })


@rol_required('Admin')
def admin_eliminar_tipo_ejercicio(request, pk):
    t = get_object_or_404(TipoEjercicio, pk=pk)
    if request.method == 'POST':
        t.delete()
        messages.success(request, 'Tipo eliminado.')
        return redirect('admin_tipos_ejercicio')
    return render(request, 'core/admin/confirmar_eliminar.html', {
        'objeto': t, 'tipo': 'tipo de ejercicio', 'notif_count': notif_count(request)
    })


# ─── COACH ────────────────────────────────────────────────────────────────────

@rol_required('Admin','Coach')
def coach_usuarios(request):
    asignaciones = RutinaUsuario.objects.select_related(
        'id_usuario', 'id_rutina', 'id_usuario__id_tipo_usuario'
    ).filter(id_usuario__id_tipo_usuario__rol='Usuario')
    return render(request, 'core/coach/usuarios.html', {
        'asignaciones': asignaciones, 'notif_count': notif_count(request)
    })


@rol_required('Coach')
def coach_rutinas(request):
    rutinas = Rutina.objects.select_related('id_nivel').order_by('nombre')
    return render(request, 'core/coach/rutinas.html', {
        'rutinas': rutinas, 'notif_count': notif_count(request)
    })


@rol_required('Coach')
def coach_crear_rutina(request):
    niveles = Nivel.objects.all()
    if request.method == 'POST':
        try:
            Rutina.objects.create(
                nombre=request.POST['nombre'],
                descripcion=request.POST.get('descripcion', ''),
                id_nivel=Nivel.objects.get(pk=request.POST['id_nivel']) if request.POST.get('id_nivel') else None,
                disciplina=request.POST.get('disciplina', ''),
                duracion_total=request.POST.get('duracion_total', ''),
                comentario=request.POST.get('comentario', ''),
            )
            messages.success(request, 'Rutina creada.')
            return redirect('coach_rutinas')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/coach/rutina_form.html', {
        'niveles': niveles, 'accion': 'Crear', 'notif_count': notif_count(request)
    })


@rol_required('Coach')
def coach_editar_rutina(request, pk):
    r = get_object_or_404(Rutina, pk=pk)
    niveles = Nivel.objects.all()
    if request.method == 'POST':
        try:
            r.nombre = request.POST['nombre']
            r.descripcion = request.POST.get('descripcion', '')
            r.id_nivel = Nivel.objects.get(pk=request.POST['id_nivel']) if request.POST.get('id_nivel') else None
            r.disciplina = request.POST.get('disciplina', '')
            r.duracion_total = request.POST.get('duracion_total', '')
            r.comentario = request.POST.get('comentario', '')
            r.save()
            messages.success(request, 'Rutina actualizada.')
            return redirect('coach_rutinas')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/coach/rutina_form.html', {
        'obj': r, 'niveles': niveles, 'accion': 'Editar', 'notif_count': notif_count(request)
    })


@rol_required('Coach')
def coach_eliminar_rutina(request, pk):
    r = get_object_or_404(Rutina, pk=pk)
    if request.method == 'POST':
        r.delete()
        messages.success(request, 'Rutina eliminada.')
        return redirect('coach_rutinas')
    return render(request, 'core/coach/confirmar_eliminar.html', {
        'objeto': r, 'tipo': 'rutina', 'notif_count': notif_count(request)
    })


@rol_required('Coach')
def coach_asignar_rutina(request, pk):
    rutina = get_object_or_404(Rutina, pk=pk)
    usuarios = Usuario.objects.filter(id_tipo_usuario__rol='Usuario')

    if request.method == 'POST':
        try:
            uid = request.POST['id_usuario']
            usuario = Usuario.objects.get(pk=uid)

            dias = request.POST.getlist('dias')
            adaptaciones = request.POST.get('adaptaciones', '').strip()

            # 1️⃣ Crear asignación de rutina
            rutina_usuario = RutinaUsuario.objects.create(
                id_usuario=usuario,
                id_rutina=rutina,
                fecha_inicio=request.POST.get('fecha_inicio') or None,
                fecha_final=request.POST.get('fecha_final') or None,
                estado='Activa',
                adaptaciones_personalizadas=adaptaciones
            )

            # 2️⃣ Crear los días asignados
            for dia in dias:
                RutinaUsuarioDia.objects.create(
                    id_rutina_usuario=rutina_usuario,
                    dia_semana=dia
                )

            # 3️⃣ Crear la notificación (¡ANTES del redirect y bien alineado!)
            crear_notificacion(
                 usuario,
                 f' Se te asignó la rutina "{rutina.nombre}"'
                )

            messages.success(request, 'Rutina asignada correctamente.')
            return redirect('coach_rutinas')

        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'core/coach/asignar_rutina.html', {
        'rutina': rutina,
        'usuarios': usuarios,
        'notif_count': notif_count(request)
    })



@rol_required('Admin','Coach','Usuario')
def coach_ejercicios(request):
    q = request.GET.get('q', '')
    ejercicios = Ejercicio.objects.select_related('tipo_ejercicio')
    if q:
        ejercicios = ejercicios.filter(nombre__icontains=q)

    return render(request, 'core/coach/ejercicios.html', {
        'ejercicios': ejercicios.order_by('nombre'),
        'q': q,
        'rol': request.session.get('rol'),   # 👈 importante
        'notif_count': notif_count(request)
    })


from .models import VideoEjercicio

@rol_required('Admin','Coach')
def coach_crear_ejercicio(request):
    tipos = TipoEjercicio.objects.all()

    if request.method == 'POST':
        try:
            ejercicio = Ejercicio.objects.create(
                nombre=request.POST['nombre'],
                tipo_ejercicio=TipoEjercicio.objects.get(pk=request.POST['tipo_ejercicio']) if request.POST.get('tipo_ejercicio') else None,
                equipo_necesario=request.POST.get('equipo_necesario', ''),
                nivel_dificultad=request.POST['nivel_dificultad'],
                instrucciones=request.POST.get('instrucciones', ''),
            )

            # 🔥 GUARDAR VIDEOS
            for i in range(1, 4):
                url = request.POST.get(f'url_video_{i}')
                if url:
                    VideoEjercicio.objects.create(
                        ejercicio=ejercicio,
                        url=url
                    )

            messages.success(request, 'Ejercicio creado con videos.')
            return redirect('coach_ejercicios')

        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'core/ejercicios/ejercicio_form.html', {
        'tipos': tipos,
        'accion': 'Crear',
        'notif_count': notif_count(request)
    })




@rol_required('Admin','Coach')
def coach_eliminar_ejercicio(request, pk):
    ej = get_object_or_404(Ejercicio, pk=pk)
    if request.method == 'POST':
        ej.delete()
        messages.success(request, 'Ejercicio eliminado.')
        return redirect('coach_ejercicios')
    return render(request, 'core/coach/confirmar_eliminar.html', {
        'objeto': ej, 'tipo': 'ejercicio', 'notif_count': notif_count(request)
    })


@rol_required('Admin','Coach')
def coach_tipos_ejercicio(request):
    tipos = TipoEjercicio.objects.all()
    return render(request, 'core/coach/tipos_ejercicio.html', {
        'tipos': tipos, 'notif_count': notif_count(request)
    })


@rol_required('Admin','Coach')
def coach_crear_tipo_ejercicio(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if nombre:
            TipoEjercicio.objects.create(nombre=nombre)
            messages.success(request, 'Tipo creado.')
            return redirect('coach_tipos_ejercicio')
    return render(request, 'core/coach/tipo_ejercicio_form.html', {
        'accion': 'Crear', 'notif_count': notif_count(request)
    })


@rol_required('Admin','Coach')
def coach_editar_tipo_ejercicio(request, pk):
    t = get_object_or_404(TipoEjercicio, pk=pk)
    if request.method == 'POST':
        t.nombre = request.POST.get('nombre', t.nombre)
        t.save()
        messages.success(request, 'Tipo actualizado.')
        return redirect('coach_tipos_ejercicio')
    return render(request, 'core/coach/tipo_ejercicio_form.html', {
        'obj': t, 'accion': 'Editar', 'notif_count': notif_count(request)
    })


@rol_required('Admin','Coach')
def coach_eliminar_tipo_ejercicio(request, pk):
    t = get_object_or_404(TipoEjercicio, pk=pk)
    if request.method == 'POST':
        t.delete()
        messages.success(request, 'Tipo eliminado.')
        return redirect('coach_tipos_ejercicio')
    return render(request, 'core/coach/confirmar_eliminar.html', {
        'objeto': t, 'tipo': 'tipo de ejercicio', 'notif_count': notif_count(request)
    })


# ─── USUARIO ──────────────────────────────────────────────────────────────────

@rol_required('Usuario')
def usuario_rutina(request):
    uid = request.session['usuario_id']

    # Rutinas asignadas al usuario
    rutinas_usuario = RutinaUsuario.objects.filter(
        id_usuario_id=uid
    ).select_related('id_rutina', 'id_rutina__id_nivel')

    # 🔥 NUEVO: días asignados a cada rutina del usuario
    dias_asignados = RutinaUsuarioDia.objects.filter(
        id_rutina_usuario__in=rutinas_usuario
    ).select_related(
        'id_rutina_usuario',
        'id_rutina_usuario__id_rutina',
        'id_rutina_usuario__id_rutina__id_nivel'
    )

    # Orden real de la semana
    orden_semana = [
    'Lunes',
    'Martes',
    'Miércoles',
    'Jueves',
    'Viernes',
    'Sábado',
    'Domingo'
]

    semana = {dia: None for dia in orden_semana}

    # 🔥 Construimos calendario semanal
    for asignacion_dia in dias_asignados:
        ru = asignacion_dia.id_rutina_usuario
        rutina = ru.id_rutina

        ejercicios = RutinaEjercicio.objects.filter(
            id_rutina=rutina
        ).select_related('id_ejercicio').order_by('orden')

        semana[asignacion_dia.dia_semana] = {
            'rutina_usuario': ru,
            'rutina': rutina,
            'ejercicios': ejercicios
        }

    return render(request, 'core/usuario/mi_rutina.html', {
        'semana_rutina': semana,
        'orden_semana': orden_semana,
        'notif_count': notif_count(request)
    })



@rol_required('Admin', 'Coach')
def agregar_ejercicio_rutina(request, rutina_id):
    rutina = get_object_or_404(Rutina, pk=rutina_id)
    ejercicios = Ejercicio.objects.all().order_by('nombre')

    if request.method == 'POST':
        try:
            ejercicio = Ejercicio.objects.get(pk=request.POST['id_ejercicio'])

            orden = request.POST.get('orden') or None

            # VALIDAR ORDEN REPETIDO EN LA MISMA RUTINA
            if orden:
                existe = RutinaEjercicio.objects.filter(
                    id_rutina=rutina,
                    orden=orden
                ).exists()

                if existe:
                    messages.error(
                        request,
                        f'El orden {orden} ya está ocupado en esta rutina.'
                    )
                    return redirect(
                        'agregar_ejercicio_rutina',
                        rutina_id=rutina.id_rutina
                    )

            # CREAR EJERCICIO EN LA RUTINA
            RutinaEjercicio.objects.create(
                id_rutina=rutina,
                id_ejercicio=ejercicio,
                series=request.POST.get('series') or None,
                repeticiones=request.POST.get('repeticiones') or None,
                dia_semana='Lunes',
                orden=orden,
                descanso=request.POST.get('descanso', '')
            )

            messages.success(request, 'Ejercicio agregado correctamente.')
            return redirect(
                'agregar_ejercicio_rutina',
                rutina_id=rutina.id_rutina
            )

        except Exception as e:
            messages.error(request, f'Error: {e}')

    rutina_ejercicios = RutinaEjercicio.objects.filter(
        id_rutina=rutina
    ).select_related('id_ejercicio').order_by('orden')

    return render(request, 'core/agregar_ejercicio_rutina.html', {
        'rutina': rutina,
        'ejercicios': ejercicios,
        'rutina_ejercicios': rutina_ejercicios,
        'notif_count': notif_count(request)
    })


@rol_required('Usuario')
def usuario_progreso(request):

    uid = request.session['usuario_id']
    usuario = Usuario.objects.get(pk=uid)

    seguimientos = Seguimiento.objects.filter(
        id_usuario_id=uid
    ).select_related('id_ejercicio').order_by('-fecha')

    # =============================
    # ESTADÍSTICAS
    # =============================

    total_sesiones = SesionEntrenamiento.objects.filter(
        id_rutina_usuario__id_usuario=usuario,
        completada=True
    ).count()

    total_ejercicios = EjercicioSesion.objects.filter(
        id_sesion__id_rutina_usuario__id_usuario=usuario,
        completado=True
    ).count()

    tiempo_total = SesionEntrenamiento.objects.filter(
        id_rutina_usuario__id_usuario=usuario,
        completada=True
    ).aggregate(
        total=Sum('duracion_segundos')
    )['total'] or 0

    horas = tiempo_total // 3600
    minutos = (tiempo_total % 3600) // 60

    # =============================
    # TOP EJERCICIOS
    # =============================

    top_ejercicios = (
        EjercicioSesion.objects
        .filter(
            id_sesion__id_rutina_usuario__id_usuario=usuario,
            completado=True
        )
        .values('id_rutina_ejercicio__id_ejercicio__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    # =============================
    # LOGROS
    # =============================

    logros_desbloqueados = UsuarioLogro.objects.filter(
        id_usuario=usuario
    ).count()

    # =============================
    # PRÓXIMO LOGRO
    # =============================

    siguiente_logro = None

    if total_sesiones < 1:
        siguiente_logro = {
            'nombre': 'Primer Paso',
            'actual': total_sesiones,
            'meta': 1
        }

    elif total_sesiones < 3:
        siguiente_logro = {
            'nombre': 'Constancia Inicial',
            'actual': total_sesiones,
            'meta': 3
        }

    elif total_sesiones < 25:
        siguiente_logro = {
            'nombre': 'Guerrero Fitness',
            'actual': total_sesiones,
            'meta': 25
        }

    elif total_sesiones < 100:
        siguiente_logro = {
            'nombre': 'Leyenda FlexFit',
            'actual': total_sesiones,
            'meta': 100
        }

    return render(request, 'core/usuario/progreso.html', {
        'seguimientos': seguimientos,
        'total_sesiones': total_sesiones,
        'total_ejercicios': total_ejercicios,
        'horas': horas,
        'minutos': minutos,
        'top_ejercicios': top_ejercicios,
        'logros_desbloqueados': logros_desbloqueados,
        'siguiente_logro': siguiente_logro,
        'notif_count': notif_count(request)
    })


@rol_required('Usuario')
def usuario_agregar_progreso(request):
    uid = request.session['usuario_id']
    ejercicios = Ejercicio.objects.all()
    if request.method == 'POST':
        try:
            Seguimiento.objects.create(
                id_usuario=Usuario.objects.get(pk=uid),
                id_ejercicio=Ejercicio.objects.get(pk=request.POST['id_ejercicio']),
                fecha=request.POST['fecha'],
                repeticiones_realizadas=request.POST.get('repeticiones_realizadas') or None,
                series_realizadas=request.POST.get('series_realizadas') or None,
                peso_usado=request.POST.get('peso_usado') or None,
                comentarios=request.POST.get('comentarios', '')
            )
            messages.success(request, 'Progreso registrado.')
            return redirect('usuario_progreso')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'core/usuario/agregar_progreso.html', {
        'ejercicios': ejercicios, 'notif_count': notif_count(request)
    })


@rol_required('Usuario')
def usuario_logros(request):
    uid = request.session['usuario_id']
    mis_logros = UsuarioLogro.objects.filter(
        id_usuario_id=uid
    ).select_related('id_logro')
    todos_logros = Logro.objects.all()
    ids_obtenidos = set(ul.id_logro_id for ul in mis_logros)
    return render(request, 'core/usuario/logros.html', {
        'mis_logros': mis_logros,
        'todos_logros': todos_logros,
        'ids_obtenidos': ids_obtenidos,
        'notif_count': notif_count(request)
    })


@rol_required('Usuario')
def usuario_explorar_rutinas(request):
    rutinas = Rutina.objects.select_related('id_nivel').order_by('nombre')
    return render(request, 'core/usuario/explorar_rutinas.html', {
        'rutinas': rutinas, 'notif_count': notif_count(request)
    })


@rol_required('Usuario')
def usuario_seleccionar_rutina(request, pk):

    rutina = get_object_or_404(Rutina, pk=pk)
    uid = request.session['usuario_id']

    if request.method == 'POST':

        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_final = request.POST.get('fecha_final')
        dias = request.POST.getlist('dias')

        ya_tiene = RutinaUsuario.objects.filter(
            id_usuario_id=uid,
            id_rutina=rutina,
            estado='Activa'
        ).exists()

        if ya_tiene:
            messages.warning(
                request,
                'Ya tienes esta rutina activa.'
            )
            return redirect('usuario_rutina')

        if not dias:
            messages.error(
                request,
                'Debes seleccionar al menos un día.'
            )
            return redirect(
                'usuario_seleccionar_rutina',
                pk=rutina.id_rutina
            )

        rutina_usuario = RutinaUsuario.objects.create(
            id_usuario=Usuario.objects.get(pk=uid),
            id_rutina=rutina,
            fecha_inicio=fecha_inicio or None,
            fecha_final=fecha_final or None,
            estado='Activa'
        )

        for dia in dias:
            RutinaUsuarioDia.objects.create(
                id_rutina_usuario=rutina_usuario,
                dia_semana=dia
            )

        messages.success(
            request,
            f'Rutina "{rutina.nombre}" añadida correctamente.'
        )

        return redirect('usuario_rutina')

    return render(
        request,
        'core/usuario/seleccionar_rutina.html',
        {
            'rutina': rutina,
            'notif_count': notif_count(request)
        }
    )

# ─── Notificaciones ───────────────────────────────────────────────────────────

@login_required_custom
def notificaciones_view(request):

    uid = request.session['usuario_id']

    # Marcar todas como leídas
    Notificacion.objects.filter(
        id_usuario_id=uid,
        estado='No Leída'
    ).update(
        estado='Leída'
    )

    notifs = (
        Notificacion.objects
        .filter(id_usuario_id=uid)
        .order_by('-fecha_hora')[:30]
    )

    return render(
        request,
        'core/notificaciones.html',
        {
            'notificaciones': notifs,
            'notif_count': 0
        }
    )

@login_required_custom
def marcar_leida(request, pk):

    n = get_object_or_404(
        Notificacion,
        pk=pk,
        id_usuario_id=request.session['usuario_id']
    )

    n.estado = 'Leída'
    n.save()

    return redirect('notificaciones')


@login_required_custom
def marcar_todas_leidas(request):

    Notificacion.objects.filter(
        id_usuario_id=request.session['usuario_id'],
        estado='No Leída'
    ).update(estado='Leída')

    return redirect('notificaciones')

# ─── Perfil ───────────────────────────────────────────────────────────────────

@login_required_custom
def perfil_view(request):
    user = get_object_or_404(
        Usuario,
        pk=request.session['usuario_id']
    )

    if request.method == 'POST':

        try:

            nombre = request.POST.get(
                'nombre',
                user.nombre
            ).strip()

            apellido = request.POST.get(
                'apellido',
                user.apellido
            ).strip()

            email = request.POST.get(
                'email',
                user.email
            ).strip()

            objetivo = request.POST.get(
                'objetivo',
                user.objetivo or ''
            ).strip()

            disciplina = request.POST.get(
                'disciplina_preferida',
                user.disciplina_preferida or ''
            ).strip()

            # VALIDACIONES

            if len(nombre) < 2:

                messages.error(
                    request,
                    'El nombre debe tener al menos 2 caracteres.'
                )

                return redirect('perfil')

            if len(apellido) < 2:

                messages.error(
                    request,
                    'El apellido debe tener al menos 2 caracteres.'
                )

                return redirect('perfil')

            if len(objetivo) > 100:

                messages.error(
                    request,
                    'El objetivo es demasiado largo.'
                )

                return redirect('perfil')

            if Usuario.objects.filter(
                email=email
            ).exclude(
                pk=user.pk
            ).exists():

                messages.error(
                    request,
                    'Ese correo ya está registrado.'
                )

                return redirect('perfil')

            # ACTUALIZAR DATOS

            user.nombre = nombre
            user.apellido = apellido
            user.email = email
            user.objetivo = objetivo
            user.disciplina_preferida = disciplina

            # FOTO DE PERFIL

            if request.FILES.get('foto_perfil'):

                foto = request.FILES['foto_perfil']

                extensiones_validas = [
                    '.jpg',
                    '.jpeg',
                    '.png',
                    '.webp'
                ]

                extension = os.path.splitext(
                    foto.name
                )[1].lower()

                if extension not in extensiones_validas:

                    messages.error(
                        request,
                        'Formato de imagen no permitido.'
                    )

                    return redirect('perfil')

                if foto.size > 2 * 1024 * 1024:

                    messages.error(
                        request,
                        'La imagen no puede superar 2 MB.'
                    )

                    return redirect('perfil')

                carpeta = os.path.join(
                    settings.MEDIA_ROOT,
                    'perfiles'
                )
                os.makedirs(
                     carpeta,
                     exist_ok=True
                )


                nombre_archivo = (
                    f"usuario_{user.id_usuario}"
                    f"{extension}"
                )

                ruta_archivo = os.path.join(
                 carpeta,
                 nombre_archivo
                )

                with open(
                    ruta_archivo,
                    'wb'
                ) as destino:

                    for chunk in foto.chunks():

                        destino.write(chunk)

                user.foto_perfil = (
                    f"perfiles/{nombre_archivo}"
                )

                print("==========================")
                print("ARCHIVO:", ruta_archivo)
                print("EXISTE:", os.path.exists(ruta_archivo))
                print("==========================")


            # CAMBIO DE CONTRASEÑA

            password_actual = request.POST.get(
                'contrasena_actual'
            )

            password_nueva = request.POST.get(
                'contrasena'
            )

            password_confirmar = request.POST.get(
                'contrasena2'
            )

            if password_nueva:

                if not check_password(
                    password_actual,
                    user.contrasena
                ):

                    messages.error(
                        request,
                        'La contraseña actual es incorrecta.'
                    )

                    return redirect('perfil')

                if password_nueva != password_confirmar:

                    messages.error(
                        request,
                        'Las contraseñas nuevas no coinciden.'
                    )

                    return redirect('perfil')

                if len(password_nueva) < 8:

                    messages.error(
                        request,
                        'La nueva contraseña debe tener mínimo 8 caracteres.'
                    )

                    return redirect('perfil')

                user.contrasena = make_password(
                    password_nueva
                )

            user.save()

            print("MEDIA ROOT:", settings.MEDIA_ROOT)
            print("FOTO BD:", user.foto_perfil)

            ruta_prueba = os.path.join(
                  settings.MEDIA_ROOT,
                 user.foto_perfil
                )

            print("RUTA FISICA:", ruta_prueba)
            print("EXISTE:", os.path.exists(ruta_prueba))




            request.session['nombre'] = (
                f"{user.nombre} {user.apellido}"
            )

            messages.success(
                request,
                'Perfil actualizado correctamente.'
            )

        except Exception as e:

            messages.error(
                request,
                f'Error: {e}'
            )

    return render(
        request,
        'core/perfil.html',
        {
            'user_obj': user,
            'notif_count': notif_count(request)
        }
    )

# ─── ÍTEM 1: CARGA MASIVA ─────────────────────────────────────────────────────

@rol_required('Admin')
def carga_masiva(request):
    """Vista principal de carga masiva de datos (ejercicios, usuarios, rutinas)."""
    if request.method == 'POST':
        tipo = request.POST.get('tipo_carga')
        archivo = request.FILES.get('archivo')

        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return redirect('carga_masiva')

        try:
            if tipo == 'ejercicios':
                creados, errores = importar_ejercicios(archivo)
            elif tipo == 'usuarios':
                creados, errores = importar_usuarios(archivo)
            elif tipo == 'rutinas':
                creados, errores = importar_rutinas(archivo)
            else:
                messages.error(request, 'Tipo de carga no válido.')
                return redirect('carga_masiva')

            if creados:
                messages.success(request, f'✅ {creados} registro(s) importado(s) correctamente.')
            for err in errores:
                messages.warning(request, err)

            if not creados and not errores:
                messages.info(request, 'El archivo estaba vacío o no tenía datos nuevos.')

        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error procesando el archivo: {e}')

        return redirect('carga_masiva')

    return render(request, 'core/admin/carga_masiva.html', {
        'notif_count': notif_count(request)
    })


# ─── ÍTEM 2: YOUTUBE API ──────────────────────────────────────────────────────

@login_required_custom
def youtube_ejercicio(request, pk):
    ejercicio = get_object_or_404(Ejercicio, pk=pk)

    # 🔥 videos automáticos (API)
    videos_api = buscar_videos_youtube(ejercicio.nombre)

    # 🔥 función BIEN ubicada (dentro de la vista)
    def extraer_video_id(url):
        try:
            if "v=" in url:
                return url.split("v=")[1].split("&")[0]
            elif "youtu.be/" in url:
                return url.split("youtu.be/")[1].split("?")[0]
            elif "shorts/" in url:
                return url.split("shorts/")[1].split("?")[0]
        except:
            return None
        return None

    # 🔥 traer videos de BD
    videos_db = ejercicio.videos.all()

    videos_guardados = []
    for v in videos_db:
        vid_id = extraer_video_id(v.url)

        print("URL:", v.url)
        print("ID:", vid_id)

        if not vid_id:
            continue

        videos_guardados.append({
            'video_id': vid_id,
            'titulo': 'Video personalizado',
            'thumbnail': f'https://img.youtube.com/vi/{vid_id}/mqdefault.jpg',
            'url': v.url,
            'embed_url': f'https://www.youtube.com/embed/{vid_id}',
        })

    # 🔥 prioridad: primero los guardados
    if videos_guardados:
        videos = videos_guardados
    else:
        videos = videos_api

    return render(request, 'core/youtube_ejercicio.html', {
        'ejercicio': ejercicio,
        'videos': videos,
        'notif_count': notif_count(request),
    })


#________________3 FILTROS_______________________

@rol_required('Admin')
def reportes(request):
    from django.db.models import Count, Avg, Case, When, IntegerField
    from django.db.models.functions import TruncMonth

    # ── Filtros ─────────────────────────────────────────────
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    nivel_filtro = request.GET.get('nivel', '')
    tipo_filtro = request.GET.get('tipo', '')

    # ── Estadísticas generales ─────────────────────────────
    total_usuarios = Usuario.objects.filter(id_tipo_usuario__rol='Usuario').count()
    total_coaches = Usuario.objects.filter(id_tipo_usuario__rol='Coach').count()
    total_ejercicios = Ejercicio.objects.count()
    total_rutinas = Rutina.objects.count()

    # ── BASE REAL (EjercicioSesion) ─────────────────────────
    ejecuciones_qs = EjercicioSesion.objects.select_related(
        'id_sesion',
        'id_rutina_ejercicio__id_ejercicio',
        'id_rutina_ejercicio__id_rutina'
    )

    # ── FILTROS SOBRE EJECUCIONES ───────────────────────────
    if fecha_desde:
        ejecuciones_qs = ejecuciones_qs.filter(
            id_sesion__fecha_inicio__gte=fecha_desde
        )

    if fecha_hasta:
        ejecuciones_qs = ejecuciones_qs.filter(
            id_sesion__fecha_inicio__lte=fecha_hasta
        )

    if tipo_filtro:
        ejecuciones_qs = ejecuciones_qs.filter(
            id_rutina_ejercicio__id_ejercicio__tipo_ejercicio__nombre__iexact=tipo_filtro
        )

    if nivel_filtro:
        ejecuciones_qs = ejecuciones_qs.filter(
            id_rutina_ejercicio__id_rutina__id_nivel__nombre__iexact=nivel_filtro
        )

    # ── MÉTRICAS DINÁMICAS ────────────────────────────────
    total_sesiones_filtradas = ejecuciones_qs.count()

    promedio_repeticiones = ejecuciones_qs.aggregate(
        promedio=Avg(
            Case(
                When(completado=True, then=1),
                default=0,
                output_field=IntegerField()
            )
        )
    )['promedio'] or 0

    # ── TOP EJERCICIOS ────────────────────────────────────
    top_ejercicios = (
    EjercicioSesion.objects   # ✅ independiente
    .values('id_rutina_ejercicio__id_ejercicio__nombre')
    .annotate(total=Count('id'))
    .order_by('-total')[:5]
)

    # ── EJERCICIOS FILTRADOS ──────────────────────────────
    ejercicios_qs = Ejercicio.objects.all()

    if tipo_filtro:
        ejercicios_qs = ejercicios_qs.filter(
            tipo_ejercicio__nombre__iexact=tipo_filtro
        )

    if nivel_filtro:
        ejercicios_qs = ejercicios_qs.filter(
            rutinaejercicio__id_rutina__id_nivel__nombre__iexact=nivel_filtro
        ).distinct()

    total_ejercicios_filtrados = ejercicios_qs.count()

    # ── RUTINAS FILTRADAS ────────────────────────────────
    rutinas_qs = Rutina.objects.all()

    if nivel_filtro:
        rutinas_qs = rutinas_qs.filter(
            id_nivel__nombre__iexact=nivel_filtro
        )

    if tipo_filtro:
        rutinas_qs = rutinas_qs.filter(
            rutinaejercicio__id_ejercicio__tipo_ejercicio__nombre__iexact=tipo_filtro
        ).distinct()

    total_rutinas_filtradas = rutinas_qs.count()

    # ── GRÁFICAS ─────────────────────────────────────────
    usuarios_por_genero = (
        Usuario.objects.filter(id_tipo_usuario__rol='Usuario')
        .values('genero')
        .annotate(total=Count('id_usuario'))
        .order_by('genero')
    )

    ejercicios_por_nivel = (
        Ejercicio.objects.values('nivel_dificultad')
        .annotate(total=Count('id_ejercicio'))
        .order_by('nivel_dificultad')
    )

    registros_por_mes = (
        Usuario.objects.filter(id_tipo_usuario__rol='Usuario')
        .annotate(mes=TruncMonth('fecha_registro'))
        .values('mes')
        .annotate(total=Count('id_usuario'))
        .order_by('mes')
    )

    # ── RUTINAS MÁS USADAS ─────────────────────────
    rutinas_mas_usadas = (
        Rutina.objects
        .annotate(total_asignaciones=Count('rutinausuario'))
        .order_by('-total_asignaciones')[:5]
    )


    # ── OPCIONES DE FILTRO ────────────────────────────────
    niveles = Nivel.objects.all()
    tipos = TipoEjercicio.objects.all()
    total_sesiones = EjercicioSesion.objects.count()

    # ── CONTEXTO ─────────────────────────────────────────
    ctx = {
        'notif_count': notif_count(request),

        # Generales
        'total_usuarios': total_usuarios,
        'total_coaches': total_coaches,
        'total_ejercicios': total_ejercicios,
        'total_rutinas': total_rutinas,
        'total_sesiones': total_sesiones,

        # Dinámicos
        'total_sesiones_filtradas': total_sesiones_filtradas,
        'promedio_repeticiones': round(promedio_repeticiones, 2),
        'total_ejercicios_filtrados': total_ejercicios_filtrados,
        'total_rutinas_filtradas': total_rutinas_filtradas,

        # Gráficas
        'usuarios_por_genero': list(usuarios_por_genero),
        'ejercicios_por_nivel': list(ejercicios_por_nivel),
        'top_ejercicios': list(top_ejercicios),
        'registros_por_mes': list(registros_por_mes),

         # 🔥 ESTE ES CLAVE
        'rutinas_mas_usadas': rutinas_mas_usadas,

        # Filtros
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'nivel_filtro': nivel_filtro,
        'tipo_filtro': tipo_filtro,

        # Selects
        'niveles': niveles,
        'tipos': tipos,

    }
    

  

    return render(request, 'core/admin/reportes.html', ctx)

#Dividir


@rol_required('Admin', 'Coach')
def ver_rutina(request, rutina_id):
    rutina = get_object_or_404(Rutina, pk=rutina_id)

    ejercicios = RutinaEjercicio.objects.filter(
        id_rutina=rutina
    ).select_related('id_ejercicio').order_by('dia_semana', 'orden')

    return render(request, 'core/ver_rutina.html', {
        'rutina': rutina,
        'ejercicios': ejercicios,
        'notif_count': notif_count(request)
    })

@login_required_custom
def usuario_iniciar_rutina(request, pk):
    rutina_usuario = get_object_or_404(RutinaUsuario, pk=pk, id_usuario_id=request.session['usuario_id'])

    sesion = SesionEntrenamiento.objects.create(
        id_rutina_usuario=rutina_usuario
    )

    return redirect('usuario_ejecutar_rutina', sesion.id_sesion)

@login_required_custom
def usuario_ejecutar_rutina(request, pk):
    sesion = get_object_or_404(
    SesionEntrenamiento,
    pk=pk,
    id_rutina_usuario__id_usuario_id=request.session['usuario_id']
)

    ejercicios = RutinaEjercicio.objects.filter(
        id_rutina=sesion.id_rutina_usuario.id_rutina
    )

    # crear registros si no existen
    for e in ejercicios:
        EjercicioSesion.objects.get_or_create(
            id_sesion=sesion,
            id_rutina_ejercicio=e
        )

    ejercicios_sesion = EjercicioSesion.objects.filter(id_sesion=sesion)

    return render(request, 'core/usuario/ejecutar_rutina.html', {
        'sesion': sesion,
        'ejercicios': ejercicios_sesion
    })

@login_required_custom
@require_POST
def completar_ejercicio(request, pk):
    ejercicio = get_object_or_404(
    EjercicioSesion,
    pk=pk,
    id_sesion__id_rutina_usuario__id_usuario_id=request.session['usuario_id']
)
    ejercicio.completado = True
    ejercicio.save()
    return JsonResponse({'ok': True})



# ─── NOTIFICACIONES ─────────────────────────────

def crear_notificacion(usuario, mensaje):
    Notificacion.objects.create(
        id_usuario=usuario,
        mensaje=mensaje,
        estado='No Leída'
    )



# ─── LOGROS ─────────────────────────────

def verificar_logros(usuario):

    total_sesiones = SesionEntrenamiento.objects.filter(
        id_rutina_usuario__id_usuario=usuario,
        completada=True
    ).count()

    logros = [
        (1, 'Primer Paso'),
        (3, 'Constancia Inicial'),
        (25, 'Guerrero Fitness'),
        (100, 'Leyenda FlexFit'),
    ]

    for meta, nombre in logros:

        if total_sesiones >= meta:

            try:
                logro = Logro.objects.get(nombre=nombre)

                ya_existe = UsuarioLogro.objects.filter(
                    id_usuario=usuario,
                    id_logro=logro
                ).exists()

                if not ya_existe:

                    UsuarioLogro.objects.create(
                        id_usuario=usuario,
                        id_logro=logro,
                        fecha_obtenida=timezone.now()
                    )

                    # 🔔 Notificación logro
                    crear_notificacion(
                        usuario,
                        f' Has desbloqueado el logro "{logro.nombre}"'
                    )

            except Logro.DoesNotExist:
                pass


@login_required_custom
@require_POST
def finalizar_rutina(request, pk, tiempo):

    sesion = get_object_or_404(
        SesionEntrenamiento,
        pk=pk,
        id_rutina_usuario__id_usuario_id=request.session['usuario_id']
    )

    # Actualizamos los campos
    sesion.fecha_fin = timezone.now()
    sesion.duracion_segundos = int(tiempo)
    sesion.completada = True
    sesion.save()

    # Usuario
    usuario = sesion.id_rutina_usuario.id_usuario

    # 🔔 Notificación rutina completada
    crear_notificacion(
        usuario,
        'Has completado una sesión de entrenamiento '
    )

    # 🏆 Verificar logros
    verificar_logros(usuario)

    return JsonResponse({'ok': True})



# ─────────────────────────────
# NOTIFICACIONES
# ─────────────────────────────

@login_required_custom
def notificaciones(request):

    usuario_id = request.session['usuario_id']

    # Obtener notificaciones
    notifs = Notificacion.objects.filter(
        id_usuario_id=usuario_id
    ).order_by('-fecha_hora')

    # Marcar como leídas
    Notificacion.objects.filter(
        id_usuario_id=usuario_id,
        estado='No Leída'
    ).update(estado='Leída')

    return render(request, 'core/notificaciones.html', {
        'notificaciones': notifs,
        'notif_count': notif_count(request)
    })


def media_view(request, path):
    
    file_path = os.path.join(settings.MEDIA_ROOT, path)

    if os.path.exists(file_path):
        return FileResponse(
            open(file_path, 'rb')
        )

    raise Http404("Archivo no encontrado")